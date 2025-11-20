import streamlit as st
import pandas as pd
import io
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Dashboard Empresarial", layout="centered")

if 'pagina' not in st.session_state:
    st.session_state['pagina'] = 'inicio'

# --- PÁGINA INICIAL ELEGANTE ---
if st.session_state['pagina'] == 'inicio':
    st.title("📊 Painel de Comparação de Gastos Empresariais")
     # Bloco de créditos do projeto/programadores logo após o título
    with st.expander(" Sobre o projeto "):
        st.markdown("""
**Projeto de Extensão:**  
Tópicos de Big Data em Python  
Faculdade Faci Wyden

**Programadores:**
- Rafael Moraes
- Bianca Santos
- Andreina Gomes

Este sistema foi desenvolvido como parte das atividades práticas e extensionistas do curso, com foco em análise comparativa de gastos empresariais via Python.
""", unsafe_allow_html=True)

    
    st.markdown("---")

    st.markdown(
        "<h4 style='text-align:center;margin-bottom:8px;'>Como usar o sistema:</h4>",
        unsafe_allow_html=True
    )

    st.markdown("""
<ol>
<li><b>Baixe o modelo</b> de planilha Excel (botão logo abaixo).</li>
<li><b>Preencha</b> somente a linha 2 com os dados da sua empresa (veja instruções abaixo).</li>
<li><b>Vá para o Dashboard</b> e faça o upload do arquivo preenchido.</li>
<li><b>Veja os gráficos, análises e exporte o relatório PDF.</b></li>
</ol>
""", unsafe_allow_html=True)

    st.markdown("---")
    with st.expander(" Como preencher a planilha Excel (com exemplo)"):
        st.markdown("""
- **Só edite a LINHA 2 (empresa). Não mexa nos títulos.**
- Use apenas NÚMEROS (sem pontos, vírgulas ou R$).
- Todos os campos devem ser preenchidos (use 0 se não tiver valor).

**Exemplo:**
| empresa        | setor   | energia | agua | ...      |
|----------------|---------|---------|------|----------|
| Minha Empresa  | Varejo  | 2500    | 800  | ...      |
""")
        st.info("Colunas: empresa (nome), setor (categoria), gastos em reais – SEM símbolo/R$!")

    st.markdown("---")
    st.markdown("###  Baixe o modelo Excel:")

    # --- Botão baixar Excel modelo ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo"
    headers = [
        "empresa", "setor", "energia", "agua", "custo_por_funcionario", "internet",
        "aluguel", "telefone", "impostos", "transporte", "marketing", "manutencao",
        "salarios", "seguranca", "limpeza"
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border_dados = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = 0
        cell.border = border_dados
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.number_format = '#,##0'
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 14
    ws.column_dimensions['A'].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    st.download_button(
        label="⬇️ Baixar Modelo Excel (.xlsx)",
        data=buffer,
        file_name="modelo_dados.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    st.markdown("---")
    st.success("💡 Dica: Não mexa em nada relacionado ao setor, o sistema já traz a referência correta!")

    st.markdown("---")
    st.markdown(
        "<div style='text-align:center;'>"
        "<b style='font-size:18px;'>Quando terminar o preenchimento, vá para o Dashboard:</b><br><br>"
        "</div>",
        unsafe_allow_html=True
    )
    st.write("")
    btn_col1, btn_col2, btn_col3 = st.columns([3,5,3])
    with btn_col2:
        if st.button("🟢 Acessar Dashboard", use_container_width=True):
            st.session_state['pagina'] = 'dashboard'
            st.rerun()

# -------------------- #
# PÁGINA DASHBOARD
elif st.session_state['pagina'] == 'dashboard':
    import plotly.graph_objects as go
    import os
    from datetime import datetime
    import tempfile
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors

    st.title("Painel de Comparação de Gastos Empresariais")
    if st.button("← Voltar para Início"):
        st.session_state['pagina'] = 'inicio'
        st.rerun()
    st.markdown("---")
    if 'importado' not in st.session_state:
        st.session_state['importado'] = False

    if not st.session_state['importado']:
        st.header("1. Faça upload do banco de empresas")
        st.markdown("Envie o arquivo Excel preenchido pela sua empresa:")
        arquivo_emp = st.file_uploader("Banco de empresas (.csv ou .xlsx)", type=["csv", "xlsx"], key='emp')

        try:
            df_setor = pd.read_excel('setores.xlsx')
            df_setor.columns = [col.lower() for col in df_setor.columns]
        except FileNotFoundError:
            st.error("❌ Erro: O arquivo 'setores.xlsx' não foi encontrado na pasta do projeto!")
            st.stop()

        if arquivo_emp:
            if arquivo_emp.name.endswith(".csv"):
                df_empresas = pd.read_csv(arquivo_emp)
            else:
                df_empresas = pd.read_excel(arquivo_emp)
            df_empresas.columns = [col.lower() for col in df_empresas.columns]
            st.session_state['df_empresas'] = df_empresas
            st.session_state['df_setor'] = df_setor
            st.success("✅ Importação realizada! Clique para avançar.")
            if st.button("Avançar para relatórios"):
                st.session_state['importado'] = True
                st.rerun()
            st.stop()
        else:
            st.stop()

    df_empresas = st.session_state['df_empresas']
    df_setor = st.session_state['df_setor']

    st.subheader("2. Configuração da Análise")
    col1, col2 = st.columns(2)
    with col1:
        nome_responsavel = st.text_area("Responsável pela análise", height=60)
    with col2:
        observacao = st.text_area("Observações (opcional)", height=60)
    st.markdown("---")

    st.subheader("3. Selecione os Dados")
    col1, col2, col3 = st.columns(3)
    with col1:
        empresa_nome = st.selectbox("Empresa:", df_empresas.iloc[:,0].unique())
    with col2:
        setor_nome = st.selectbox("Setor de referência:", df_setor.iloc[:,0].unique())
    with col3:
        tipo_grafico = st.selectbox("Tipo de gráfico:", ["Barras Vertical", "Barras Horizontal", "Pizza"])

    empresa = df_empresas[df_empresas.iloc[:,0] == empresa_nome].iloc[0]
    setor_row = df_setor[df_setor.iloc[:,0] == setor_nome].iloc[0]
    colunas_numericas = [c for c in df_setor.columns if c != df_setor.columns[0]]
    
    def formatar_nome_indicador(nome):
        if nome == 'custo_por_funcionario':
            return 'Salarios'
        return nome.replace("_", " ").capitalize()

    indicadores_grafico = [formatar_nome_indicador(c) for c in colunas_numericas]
    valores_empresa_grafico = [empresa[c] for c in colunas_numericas]
    valores_setor_grafico = [setor_row[c] for c in colunas_numericas]

    if tipo_grafico == "Barras Vertical":
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=indicadores_grafico,
            y=valores_setor_grafico,
            name=f"Média {setor_nome}",
            marker_color="#2477EA",
            text=[f'{int(v):,}'.replace(',', '.') for v in valores_setor_grafico],
            textposition='outside'
        ))
        fig.add_trace(go.Bar(
            x=indicadores_grafico,
            y=valores_empresa_grafico,
            name=empresa_nome,
            marker_color="#8ECEED",
            text=[f'{int(v):,}'.replace(',', '.') for v in valores_empresa_grafico],
            textposition='outside'
            
            
        ))
        fig.update_layout(
    barmode='group',
    title=None,
    xaxis=dict(title='Indicador', tickangle=-30, automargin=True, showgrid=False),
    yaxis=dict(
        title='Valor (R$)',
        showgrid=True,
        nticks=7,
        type='log'       # linha liagrosaaaaa
    ),
    legend=dict(title='Referências', orientation='h', yanchor='bottom', y=1.08, xanchor='center', x=0.5),
    height=700,
    width=1200,
    margin=dict(l=42, r=32, t=50, b=60)
)

    elif tipo_grafico == "Barras Horizontal":
        fig = go.Figure()
        fig.add_trace(go.Bar(
            y=indicadores_grafico,
            x=valores_setor_grafico,
            name=f"Média {setor_nome}",
            orientation='h',
            marker_color="#2477EA",
            text=[f'{int(v):,}'.replace(',', '.') for v in valores_setor_grafico],
            textposition='outside'
        ))
        fig.add_trace(go.Bar(
            y=indicadores_grafico,
            x=valores_empresa_grafico,
            name=empresa_nome,
            orientation='h',
            marker_color="#8ECEED",
            text=[f'{int(v):,}'.replace(',', '.') for v in valores_empresa_grafico],
            textposition='outside'
        ))
        fig.update_layout(
            barmode='group',
            title=None,
            xaxis=dict(title='Valor (R$)', showgrid=True, nticks=7),
            yaxis=dict(title='Indicador', automargin=True, showgrid=False),
            legend=dict(title='Referências', orientation='h', yanchor='bottom', y=1.08, xanchor='center', x=0.5),
            height=650,
            margin=dict(l=120, r=32, t=40, b=70)
        )

    elif tipo_grafico == "Pizza":
        pie_colors = ["#2477EA", "#8ECEED", "#A0C4F6", "#B9D5F8", "#46628d", "#83ACE7", "#69b4fa"]
        fig = go.Figure()
        fig.add_trace(go.Pie(
            labels=indicadores_grafico,
            values=valores_empresa_grafico,
            name=empresa_nome,
            hole=0.3,
            marker=dict(colors=pie_colors)
        ))
        fig.update_traces(textinfo='label+percent')
        fig.update_layout(title=None)

    st.subheader("4. Visualização Gráfica")
    st.plotly_chart(fig, use_container_width=True)
    st.markdown("---")

    tabela = []
    for i, c in enumerate(colunas_numericas):
        val = empresa[c]
        media = setor_row[c]
        diff = (val - media) / media if media else 0
        if diff < -0.1:
            situacao = "Abaixo"
        elif diff > 0.1:
            situacao = "Acima"
        else:
            situacao = "Na média"
        tabela.append([formatar_nome_indicador(c), val, media, situacao])

    st.subheader("5. Comparação Detalhada")
    st.dataframe(pd.DataFrame(tabela, columns=["Gasto", "Empresa", "Média Setor", "Situação"]), use_container_width=True)
    st.markdown("---")

    total_empresa = sum(valores_empresa_grafico)
    total_media = sum(valores_setor_grafico)
    percentual = ((total_empresa - total_media) / total_media) * 100 if total_media else 0
    if percentual < -10:
        resumo_executivo = f"A empresa **{empresa_nome}** gasta **{abs(percentual):.1f}% MENOS** que a média do setor."
    elif percentual > 10:
        resumo_executivo = f"A empresa **{empresa_nome}** gasta **{percentual:.1f}% MAIS** que a média do setor."
    else:
        resumo_executivo = f"A empresa **{empresa_nome}** gasta dentro da média do setor (**{percentual:+.1f}%**)."

    analise_detalhada = ""
    principais_acima = [row for row in tabela if row[3] == "Acima"]
    principais_abaixo = [row for row in tabela if row[3] == "Abaixo"]
    if principais_acima:
        analise_detalhada += "**Acima da média:** " + ", ".join(f"`{row[0]}`" for row in principais_acima) + ".\n"
    if principais_abaixo:
        analise_detalhada += "**Abaixo da média:** " + ", ".join(f"`{row[0]}`" for row in principais_abaixo) + "."
    if not (principais_acima or principais_abaixo):
        analise_detalhada += "Os gastos estão próximos da média em todos os principais indicadores."

    st.subheader("6. Resumo Executivo")
    st.info(resumo_executivo)
    st.markdown("### Análise Detalhada")
    st.success(analise_detalhada if analise_detalhada else "Nenhum destaque para cima ou para baixo.")
    if observacao.strip():
        st.markdown("### Observações")
        st.warning(observacao)

    def salvar_grafico_png_matplotlib(indicadores, valores_empresa, valores_setor, empresa_nome, setor_nome, tipo_grafico):
        fig_mpl, ax = plt.subplots(figsize=(12, 5), dpi=100)
        indices = np.arange(len(indicadores))
        width = 0.36

        if tipo_grafico == "Barras Vertical":
            ax.bar(indices - width/2, valores_setor, width, label=f'Média {setor_nome}', color="#2477EA", alpha=0.9)
            ax.bar(indices + width/2, valores_empresa, width, label=empresa_nome, color="#8ECEED", alpha=0.9)
            ax.set_xticks(indices)
            ax.set_xticklabels(indicadores, rotation=30, ha="right", fontsize=10)
            ax.set_ylabel('Valor (R$)', fontsize=12)
            ax.set_title('Gráfico Comparativo de Gastos', fontsize=13, fontweight='bold')
            ax.legend(fontsize=10)
            ax.grid(axis='y', linestyle='--', alpha=0.6)
            fig_mpl.tight_layout()
        elif tipo_grafico == "Barras Horizontal":
            ax.barh(indices - width/2, valores_setor, width, label=f'Média {setor_nome}', color="#2477EA", alpha=0.9)
            ax.barh(indices + width/2, valores_empresa, width, label=empresa_nome, color="#8ECEED", alpha=0.9)
            ax.set_yticks(indices)
            ax.set_yticklabels(indicadores, fontsize=10)
            ax.set_xlabel('Valor (R$)', fontsize=12)
            ax.set_title('Gráfico Comparativo de Gastos', fontsize=13, fontweight='bold')
            ax.legend(fontsize=10)
            ax.grid(axis='x', linestyle='--', alpha=0.6)
            fig_mpl.tight_layout()
        else:
            colors_pizza = ["#2477EA", "#8ECEED", "#A0C4F6", "#B9D5F8", "#46628d", "#83ACE7", "#69b4fa"]
            ax.pie(valores_empresa, labels=indicadores, autopct='%1.1f%%', colors=colors_pizza, startangle=90)
            ax.set_title(f'Gráfico de Pizza - {empresa_nome}', fontsize=13, fontweight='bold')
            ax.axis('equal')

        tmp_png = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        plt.savefig(tmp_png.name, bbox_inches='tight')
        plt.close(fig_mpl)
        return tmp_png.name

    def gerar_pdf_resumido(resumo_executivo, analise_detalhada, tabela, empresa_nome, setor_nome, nome_responsavel, observacao, indicadores, valores_empresa, valores_setor, tipo_grafico):
        imgpath = salvar_grafico_png_matplotlib(indicadores, valores_empresa, valores_setor, empresa_nome, setor_nome, tipo_grafico)
        tmpfile = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        doc = SimpleDocTemplate(tmpfile.name, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph("Resumo Comparativo", styles['Heading1']))
        story.append(Spacer(1, 10))
        datahora = datetime.now().strftime("%d/%m/%Y %H:%M")
        story.append(Paragraph(f"Data da análise: {datahora}", styles['Normal']))
        if nome_responsavel.strip():
            story.append(Paragraph(f"Responsável pela análise: {nome_responsavel}", styles['Normal']))
        story.append(Paragraph(f"Empresa analisada: {empresa_nome}", styles['Normal']))
        story.append(Paragraph(f"Setor de referência: {setor_nome}", styles['Normal']))
        if observacao.strip():
            story.append(Spacer(1, 8))
            story.append(Paragraph(f"Observações: {observacao}", styles['Normal']))
        
        story.append(Spacer(1, 15))

        story.append(Paragraph("Gráfico Comparativo", styles['Heading2']))
        story.append(Spacer(1, 8))
        story.append(RLImage(imgpath, width=450, height=225))
        story.append(Spacer(1, 12))

        head = ['Gasto', 'Empresa', 'Média Setor', 'Situação']
        dados_tab = [head] + [[str(cell) for cell in linha] for linha in tabela]
        table = Table(dados_tab, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('TEXTCOLOR',(0,0),(-1,0),colors.black),
            ('ALIGN',(1,1),(-1,-1),'CENTER'),
            ('FONTNAME', (0,0),(-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0),(-1,-1), 10),
            ('BOTTOMPADDING', (0,0),(-1,0), 6),
            ('LINEBELOW', (0,0),(-1,0), 1, colors.black),
            ('BACKGROUND', (0,1),(-1,-1), colors.whitesmoke),
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ]))
        story.append(table)
        story.append(Spacer(1, 18))

        story.append(Paragraph("Resumo Executivo", styles['Heading2']))
        story.append(Paragraph(resumo_executivo, styles['Normal']))
        story.append(Spacer(1, 12))
        story.append(Paragraph("Análise Detalhada", styles['Heading2']))
        story.append(Paragraph(analise_detalhada, styles['Normal']))

        doc.build(story)
        with open(tmpfile.name, "rb") as pdf_file:
            pdf_bytes = pdf_file.read()
        tmpfile.close()
        try:
            os.remove(imgpath)
        except Exception:
            pass
        return pdf_bytes

    pdf_bytes = gerar_pdf_resumido(
        resumo_executivo,
        analise_detalhada,
        tabela,
        empresa_nome,
        setor_nome,
        nome_responsavel,
        observacao,
        indicadores_grafico,
        valores_empresa_grafico,
        valores_setor_grafico,
        tipo_grafico
    )

    st.markdown("---")
    st.subheader("7. Exportar PDF do relatório")
    st.download_button(
        label="📥 Baixar Relatório em PDF",
        data=pdf_bytes,
        file_name=f"resumo_{empresa_nome}.pdf",
        mime="application/pdf",
        use_container_width=True
    )
